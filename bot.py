import requests
import os
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from aiogram import Bot, types, F
from aiogram.filters.command import Command
from aiogram.types import CallbackQuery
from aiogram import Dispatcher
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from aiogram.types import FSInputFile, URLInputFile, BufferedInputFile
from openpyxl import load_workbook
import asyncio
from config import TOKEN

bot = Bot(token=TOKEN)
dp = Dispatcher()

URL_RU = 'https://crc-reg.com/about/our-clients/'

r_url_ru = requests.get(URL_RU)
soup_url_ru = BeautifulSoup(r_url_ru.text, 'html.parser')
table_ru = soup_url_ru.find('table')

clients = soup_url_ru.find_all('td', class_='clients-table-col-2')
inns = soup_url_ru.find_all('td', class_='clients-table-col-3')
ogrns = soup_url_ru.find_all('td', class_='clients-table-col-4')

ao_ru = []
for client in clients:
    ao_ru.append(client.get_text(strip=True))

inn_ru = []
for inn in inns:
    inn_ru.append("; ИНН: " + inn.get_text(strip=True))

ogrn_ru = []
for ogrn in ogrns:
    ogrn_ru.append("; ОГРН: " + ogrn.get_text(strip=True) + '\n\n')

all_info = zip(ao_ru, inn_ru, ogrn_ru)
all_info_list = list(all_info)

ao_ru_str = ''.join(''.join(elems) for elems in all_info_list)

URL_CONTACTS_ru = 'https://crc-reg.com/contacts/'

r_contact_ru = requests.get(URL_CONTACTS_ru)
soup_contact_ru = BeautifulSoup(r_contact_ru.text, 'html.parser')
contacts_ru = soup_contact_ru.find_all('div', class_='contacts__sub-ttl')
contact_str_ru = ' '
for contact_ru in contacts_ru:
    contact_str_ru = contact_str_ru.join(contact_ru.text.split(' '))
contact_str_ru_replace = contact_str_ru.replace('Телеграм-бот', '')

bank_info_ru = soup_contact_ru.find_all('div', class_='banks')
for bank in bank_info_ru:
    bank_str_ru = bank.text

btn_ru = types.KeyboardButton(text='Русский язык')
btn_en = types.KeyboardButton(text='English language')

btn_list_ru = types.KeyboardButton(text='Реестры')

btn_return_A_f_ru = types.KeyboardButton(text='Вернуться к разделу "Анкеты"')
btn_return_orders_ru = types.KeyboardButton(text='Вернуться к разделу "Распоряжения"')
btn_return_p_s_d_ru = types.KeyboardButton(text='Вернуться к разделу "Порядок предоставления документов"')
btn_return_d_o_a_ru = types.KeyboardButton(text='Вернуться к разделу "Документы для открытия лицевого счета"')
btn_return_d_t_p_i_ru = types.KeyboardButton(
    text='Вернуться к разделу "Документы для внесения записи в информацию ЛС о ЗЛ"')
btn_return_price_ru = types.KeyboardButton(text='Вернуться к разделу "Прейскуранты"')

btn_main_menu_ru = types.KeyboardButton(text='В главное меню')

btn_p_f_s_ru = types.KeyboardButton(text='Порядок заполнения и предоставления анкет, распоряжений и запросов')
btn_d_c_t_ru = types.KeyboardButton(text='Документы для проведения операций')

btn_price_ru = types.KeyboardButton(text='Прейскуранты')

btn_contacts_ru = types.KeyboardButton(text='Контакты')

btn_A_f_ru = types.KeyboardButton(text='Анкеты')
btn_orders_ru = types.KeyboardButton(text='Распоряжения')
btn_p_s_d_ru = types.KeyboardButton(text='Порядок предоставления документов')
btn_d_o_a_ru = types.KeyboardButton(text='Документы для открытия лицевого счета')
btn_d_t_p_i_ru = types.KeyboardButton(text='Документы для внесения записи в информацию ЛС о ЗЛ')

inline_btn_1_2_3_49_50_51_ru = []
inline_btn_1_2_3_49_50_51_ru.append([InlineKeyboardButton(text='Прейскурант услуг Эмитентам', callback_data='btn1_ru')])
inline_btn_1_2_3_49_50_51_ru.append(
    [InlineKeyboardButton(text='Прейскурант услуг зарегистрированным лицам', callback_data='btn2_ru')])
inline_btn_1_2_3_49_50_51_ru.append(
    [InlineKeyboardButton(text='Прейскурант доп.услуг зарегистрированным лицам', callback_data='btn3_ru')])
inline_btn_1_2_3_49_50_51_ru.append(
    [InlineKeyboardButton(text='Прейскурант услуг МКПАО «Алсиб»', callback_data='btn53_ru')])
inline_btn_1_2_3_49_50_51_ru.append(
    [InlineKeyboardButton(text='Прейскурант услуг зарегистрированным лицам МКПАО «Алсиб»', callback_data='btn54_ru')])
inline_btn_1_2_3_49_50_51_ru.append([InlineKeyboardButton(
    text='Прейскурант доп.услуг зарегистрированным лицам МКПАО «Алсиб»', callback_data='btn55_ru')])
inline_kb_full_ru = types.InlineKeyboardMarkup(inline_keyboard=inline_btn_1_2_3_49_50_51_ru)

inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru = []
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Анкета Эмитента', callback_data='btn4_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета физического лица', callback_data='btn5_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета юридического лица', callback_data='btn6_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета нотариуса', callback_data='btn7_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append([InlineKeyboardButton(
    text='Заявление-анкета органа государственной власти (органа местного самоуправления)', callback_data='btn8_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета доверительного управляющего', callback_data='btn9_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета эскроу-агента', callback_data='btn10_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета инвестиционного товарищества', callback_data='btn11_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета. Цифровые финансовые активы', callback_data='btn12_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append([InlineKeyboardButton(
    text='Заявление-анкета иностранной структуры без образования юридического лица', callback_data='btn13_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета залогодержателя (физическое лицо)', callback_data='btn14_ru')])
inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru.append(
    [InlineKeyboardButton(text='Заявление-анкета залогодержателя (юридическое лицо)', callback_data='btn15_ru')])
inline_kb_full2_ru = InlineKeyboardMarkup(inline_keyboard=inline_btn_4_5_6_7_8_9_10_11_12_13_14_15_ru)

inline_btn_16_17_18_19_20_21_22_23_ru = []
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Распоряжение о совершении операции', callback_data='btn16_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Залоговое распоряжение', callback_data='btn17_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append([InlineKeyboardButton(
    text='Распоряжение о внесении изменений о заложенных ЦБ и условиях залога', callback_data='btn18_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Распоряжение о передаче прав залога', callback_data='btn19_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Распоряжение о прекращении залога', callback_data='btn20_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Распоряжение на предоставление информации из реестра', callback_data='btn21_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Распоряжение на снятие факта ограничения операций с ЦБ', callback_data='btn22_ru')])
inline_btn_16_17_18_19_20_21_22_23_ru.append(
    [InlineKeyboardButton(text='Распоряжение владельца о передаче ЦБ на депозитный ЛС', callback_data='btn23_ru')])
inline_kb_full3_ru = InlineKeyboardMarkup(inline_keyboard=inline_btn_16_17_18_19_20_21_22_23_ru)

inline_btn_24_25_ru = []
inline_btn_24_25_ru.append(
    [InlineKeyboardButton(text='Порядок предоставления документов для открытия ЛС', callback_data='btn24_ru')])
inline_btn_24_25_ru.append([InlineKeyboardButton(
    text='Порядок предоставления документов для совершения операции и\n для предоставления информации',
    callback_data='btn25_ru')])
inline_kb_full4_ru = InlineKeyboardMarkup(inline_keyboard=inline_btn_24_25_ru)

inline_btn_26_27_28_29_30_31_ru = []
inline_btn_26_27_28_29_30_31_ru.append(
    [InlineKeyboardButton(text='Документы для физического лица', callback_data='btn26_ru')])
inline_btn_26_27_28_29_30_31_ru.append([InlineKeyboardButton(text='Документы для нотариуса', callback_data='btn27_ru')])
inline_btn_26_27_28_29_30_31_ru.append(
    [InlineKeyboardButton(text='Документы ЮЛ, являющемуся резидентом РФ', callback_data='btn28_ru')])
inline_btn_26_27_28_29_30_31_ru.append([InlineKeyboardButton(
    text='Документы для органа государственной власти (органа местного самоуправления)', callback_data='btn29_ru')])
inline_btn_26_27_28_29_30_31_ru.append(
    [InlineKeyboardButton(text='Документы для ЮЛ-нерезидента', callback_data='btn30_ru')])
inline_btn_26_27_28_29_30_31_ru.append(
    [InlineKeyboardButton(text='Документы для открытия казначейского ЛС эмитенту', callback_data='btn31_ru')])
inline_kb_full5_ru = InlineKeyboardMarkup(inline_keyboard=inline_btn_26_27_28_29_30_31_ru)

inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru = []
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для физического лица', callback_data='btn32_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для юридического лица', callback_data='btn33_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для передачи ЦБ при совершении сделки', callback_data='btn34_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для передачи ЦБ при наследовании', callback_data='btn35_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для исполнения судебных актов', callback_data='btn36_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для передачи ЦБ при реорганизации', callback_data='btn37_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для передачи ЦБ при ликвидации ЮЛ', callback_data='btn38_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Документы для передачи ЦБ при приватизации', callback_data='btn39_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append([InlineKeyboardButton(
    text='Для внесения записи о фиксации/прекращении права залога (последующего залога) ЦБ', callback_data='btn40_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append([InlineKeyboardButton(
    text='Для внесения записи о факте фиксации ограничения операций с ЦБ по их полной оплате',
    callback_data='btn41_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append([InlineKeyboardButton(
    text='Для внесения записи о фиксации/снятии факта ограничения операций с ЦБ по ЛС ЗЛ', callback_data='btn42_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Для внесения записи о зачислении/списании ЦБ со счета НД', callback_data='btn43_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Для зачисления на ЛС НД заложенных ЦБ', callback_data='btn44_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Особенности проведения операций по ЛС ДУ', callback_data='btn45_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Особенности проведения операций по депозитному ЛС', callback_data='btn46_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Объединение ЛС в реестре', callback_data='btn47_ru')])
inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru.append(
    [InlineKeyboardButton(text='Закрытие ЛС', callback_data='btn48_ru')])
inline_kb_full6_ru = InlineKeyboardMarkup(
    inline_keyboard=inline_btn_32_33_34_35_36_37_38_39_40_41_42_43_44_45_46_47_48_ru)

greet_kb = ReplyKeyboardBuilder()
greet_kb.row(btn_ru, width=1)
greet_kb.row(btn_en, width=1)

greet_kb_ru = ReplyKeyboardBuilder()
greet_kb_ru.row(btn_list_ru, width=1)
greet_kb_ru.row(btn_p_f_s_ru, width=1)
greet_kb_ru.row(btn_d_c_t_ru, width=1)
greet_kb_ru.row(btn_price_ru, width=1)
greet_kb_ru.row(btn_contacts_ru, width=1)

reply_greet_3_ru = []
reply_greet_3_ru.append(btn_main_menu_ru)
greet_kb2_ru = ReplyKeyboardMarkup(keyboard=[reply_greet_3_ru], resize_keyboard=True)

greet_kb3_ru = ReplyKeyboardBuilder()
greet_kb3_ru.row(btn_A_f_ru, width=1)
greet_kb3_ru.row(btn_orders_ru, width=1)
greet_kb3_ru.row(btn_p_s_d_ru, width=1)
greet_kb3_ru.row(btn_main_menu_ru, width=1)

greet_kb4_ru = ReplyKeyboardBuilder()
greet_kb4_ru.row(btn_d_o_a_ru, width=1)
greet_kb4_ru.row(btn_d_t_p_i_ru, width=1)
greet_kb4_ru.row(btn_main_menu_ru, width=1)

greet_kb5_ru = ReplyKeyboardBuilder()
greet_kb5_ru.row(btn_return_A_f_ru, width=1)
greet_kb5_ru.row(btn_main_menu_ru, width=1)

greet_kb6_ru = ReplyKeyboardBuilder()
greet_kb6_ru.row(btn_return_orders_ru, width=1)
greet_kb6_ru.row(btn_main_menu_ru, width=1)

greet_kb7_ru = ReplyKeyboardBuilder()
greet_kb7_ru.row(btn_return_p_s_d_ru, width=1)
greet_kb7_ru.row(btn_main_menu_ru, width=1)

greet_kb8_ru = ReplyKeyboardBuilder()
greet_kb8_ru.row(btn_return_d_o_a_ru, width=1)
greet_kb8_ru.row(btn_main_menu_ru, width=1)

greet_kb9_ru = ReplyKeyboardBuilder()
greet_kb9_ru.row(btn_return_d_t_p_i_ru, width=1)
greet_kb9_ru.row(btn_main_menu_ru, width=1)

greet_kb10_ru = ReplyKeyboardBuilder()
greet_kb10_ru.row(btn_return_price_ru, width=1)
greet_kb10_ru.row(btn_main_menu_ru, width=1)

btn_return_A_f_en = types.KeyboardButton(text='Go back to the section "Application forms"')
btn_return_orders_en = types.KeyboardButton(text='Go back to the section "Orders"')
btn_return_p_s_d_en = types.KeyboardButton(text='Go back to the section "Procedure for submitting documents"')
btn_return_d_o_a_en = types.KeyboardButton(
    text='Go back to the section "Procedure for submitting documents to open a personal account"')
btn_return_d_t_p_i_en = types.KeyboardButton(
    text='Go back to the section "Procedure for submitting documents for a transaction and for providing information"')
btn_main_menu_en = types.KeyboardButton(text='Go back to the section "To the main menu"')
btn_p_f_s_en = types.KeyboardButton(
    text='Procedure for filling out and submitting Application forms, Orders and Requests')
btn_d_c_t_en = types.KeyboardButton(text='Documents for conducting transactions')

btn_contacts_en = types.KeyboardButton(text='Contacts')

btn_A_f_en = types.KeyboardButton(text='Application forms')
btn_orders_en = types.KeyboardButton(text='Orders')
btn_p_s_d_en = types.KeyboardButton(text='Procedure for submitting documents')
btn_d_o_a_en = types.KeyboardButton(text='Procedure for submitting documents to open a personal account')
btn_d_t_p_i_en = types.KeyboardButton(
    text='Procedure for submitting documents for a transaction and for providing information')

inline_btn_1_2_3_4_5_6_7_8_en = []
inline_btn_1_2_3_4_5_6_7_8_en.append([InlineKeyboardButton(text='Issuer Application Form', callback_data='btn1_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append(
    [InlineKeyboardButton(text='Individual Application Form', callback_data='btn2_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append(
    [InlineKeyboardButton(text='Legal Entity Application Form', callback_data='btn3_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append([InlineKeyboardButton(text='Notary Application Form', callback_data='btn4_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append(
    [InlineKeyboardButton(text='Authorised Body Application Form', callback_data='btn5_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append([InlineKeyboardButton(text='Trustee Application Form', callback_data='btn6_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append(
    [InlineKeyboardButton(text='Pledgee(individual) Application Form', callback_data='btn7_en')])
inline_btn_1_2_3_4_5_6_7_8_en.append(
    [InlineKeyboardButton(text='Pledgee(legal entity) Application Form', callback_data='btn8_en')])
inline_kb_full_en = InlineKeyboardMarkup(inline_keyboard=inline_btn_1_2_3_4_5_6_7_8_en)

inline_btn_9_10_11_12_13_14_15_16_en = []
inline_btn_9_10_11_12_13_14_15_16_en.append([InlineKeyboardButton(text='Transaction Order', callback_data='btn9_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append([InlineKeyboardButton(text='Pledge Order', callback_data='btn10_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append([InlineKeyboardButton(
    text='Order to amend the pledged securities and the terms and conditions of the pledge', callback_data='btn11_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append(
    [InlineKeyboardButton(text='Order to assign pledge rights', callback_data='btn12_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append(
    [InlineKeyboardButton(text='Order to terminate pledge', callback_data='btn13_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append(
    [InlineKeyboardButton(text='Order to provide information from the registry', callback_data='btn14_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append(
    [InlineKeyboardButton(text='Order to lift restrictions on transactions with securities', callback_data='btn15_en')])
inline_btn_9_10_11_12_13_14_15_16_en.append([InlineKeyboardButton(
    text="Owner's order to transfer securities to the deposit account", callback_data='btn16_en')])
inline_kb_full_2_en = InlineKeyboardMarkup(inline_keyboard=inline_btn_9_10_11_12_13_14_15_16_en)

inline_btn_17_18_en = []
inline_btn_17_18_en.append([InlineKeyboardButton(text='Procedure for submitting documents to open a personal account',
                                                 callback_data='btn17_en')])
inline_btn_17_18_en.append([InlineKeyboardButton(
    text='Procedure for submitting documents for a transaction and for providing information',
    callback_data='btn18_en')])
inline_kb_full_3_en = InlineKeyboardMarkup(inline_keyboard=inline_btn_17_18_en)

inline_btn_19_20_21_22_23_24_25_en = []
inline_btn_19_20_21_22_23_24_25_en.append(
    [InlineKeyboardButton(text='Documents for an individual', callback_data='btn19_en')])
inline_btn_19_20_21_22_23_24_25_en.append(
    [InlineKeyboardButton(text='Documents for the notary', callback_data='btn20_en')])
inline_btn_19_20_21_22_23_24_25_en.append([InlineKeyboardButton(
    text='Documents of a legal entity being a resident of the Russian Federation', callback_data='btn21_en')])
inline_btn_19_20_21_22_23_24_25_en.append(
    [InlineKeyboardButton(text='Documents for the Authorised Body', callback_data='btn22_en')])
inline_btn_19_20_21_22_23_24_25_en.append(
    [InlineKeyboardButton(text='Documents for non-resident legal entity', callback_data='btn23_en')])
inline_btn_19_20_21_22_23_24_25_en.append(
    [InlineKeyboardButton(text='Documents to open a treasury account for the issuer', callback_data='btn24_en')])
inline_btn_19_20_21_22_23_24_25_en.append([InlineKeyboardButton(
    text='Documents to open a personal account for common hared ownership', callback_data='btn25_en')])
inline_kb_full_4_en = InlineKeyboardMarkup(inline_keyboard=inline_btn_19_20_21_22_23_24_25_en)

inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en = []
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Documents for an individual', callback_data='btn26_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Documents for a legal entity', callback_data='btn27_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='Documents to transfer securities upon execution of a transaction', callback_data='btn28_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Documents to transfer securities upon inheritance', callback_data='btn29_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Documents to enforce court orders', callback_data='btn30_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Documents to transfer securities upon reorganisation', callback_data='btn31_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='Documents to transfer securities upon liquidation of a legal entity', callback_data='btn32_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Documents to transfer securities upon privatisation', callback_data='btn33_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='In order to make an entry on the recording / termination of a pledge (subsequent pledge) of securities',
    callback_data='btn34_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='To make an entry on committing restrictions on securities transactions to pay them in full',
    callback_data='btn35_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='To make an entry on committing / removing restrictions on securities transactions on the personal account of a registered person',
    callback_data='btn36_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='To make an entry on deposition / withdrawal of securities to / from an account of a nominee holder',
    callback_data='btn37_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text='To deposit pledged securities to an account of a nominee holder', callback_data='btn38_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append([InlineKeyboardButton(
    text="Specifics of transactions in the trustee's personal account", callback_data='btn39_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Specifics of deposit personal account transactions', callback_data='btn40_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Pooling of personal accounts in the register', callback_data='btn41_en')])
inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en.append(
    [InlineKeyboardButton(text='Closure of a personal account', callback_data='btn42_en')])
inline_kb_full_5_en = InlineKeyboardMarkup(
    inline_keyboard=inline_btn_26_27_28_29_30_31_32_33_34_35_36_37_38_39_40_41_42_en)

greet_kb_en = ReplyKeyboardBuilder()
greet_kb_en.row(btn_p_f_s_en, width=1)
greet_kb_en.row(btn_d_c_t_en, width=1)
greet_kb_en.row(btn_contacts_en, width=1)

greet_kb_2_en = ReplyKeyboardBuilder()
greet_kb_2_en.row(btn_main_menu_en, width=1)

greet_kb_3_en = ReplyKeyboardBuilder()
greet_kb_3_en.row(btn_A_f_en, width=1)
greet_kb_3_en.row(btn_orders_en, width=1)
greet_kb_3_en.row(btn_p_s_d_en, width=1)
greet_kb_3_en.row(btn_main_menu_en, width=1)

greet_kb_4_en = ReplyKeyboardBuilder()
greet_kb_4_en.row(btn_d_o_a_en, width=1)
greet_kb_4_en.row(btn_d_t_p_i_en, width=1)
greet_kb_4_en.row(btn_main_menu_en, width=1)

greet_kb_5_en = ReplyKeyboardBuilder()
greet_kb_5_en.row(btn_return_A_f_en, width=1)
greet_kb_5_en.row(btn_main_menu_en, width=1)

greet_kb_6_en = ReplyKeyboardBuilder()
greet_kb_6_en.row(btn_return_orders_en, width=1)
greet_kb_6_en.row(btn_main_menu_en, width=1)

greet_kb_7_en = ReplyKeyboardBuilder()
greet_kb_7_en.row(btn_return_p_s_d_en, width=1)
greet_kb_7_en.row(btn_main_menu_en, width=1)

greet_kb_8_en = ReplyKeyboardBuilder()
greet_kb_8_en.row(btn_return_d_o_a_en, width=1)
greet_kb_8_en.row(btn_main_menu_en, width=1)

greet_kb_9_en = ReplyKeyboardBuilder()
greet_kb_9_en.row(btn_return_d_t_p_i_en, width=1)
greet_kb_9_en.row(btn_main_menu_en, width=1)


@dp.message(Command('start'))
async def send_welcome(message: types.Message):
    await message.answer('Выберите язык / Select a language', reply_markup=greet_kb.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Русский язык')
async def ru_language(message: types.Message):
    await message.answer('Выбран русский язык.\n Выбери нужную кнопку в меню',
                         reply_markup=greet_kb_ru.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'В главное меню')
async def main_menu_ru(message: types.Message):
    await message.answer('Выберите нужную кнопку в меню', reply_markup=greet_kb_ru.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Вернуться к разделу "Анкеты"')
async def back_a_f_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full2_ru)


@dp.message(lambda message: message.text == 'Вернуться к разделу "Распоряжения"')
async def back_orders_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full3_ru)


@dp.message(lambda message: message.text == 'Вернуться к разделу "Порядок предоставления документов"')
async def back_p_s_d_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full4_ru)


@dp.message(lambda message: message.text == 'Вернуться к разделу "Документы для открытия лицевого счета"')
async def back_p_p_a_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full5_ru)


@dp.message(lambda message: message.text == 'Вернуться к разделу "Документы для внесения записи в информацию ЛС о ЗЛ"')
async def back_p_s_t_p_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full6_ru)


@dp.message(lambda message: message.text == 'Вернуться к разделу "Прейскуранты"')
async def back_price_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_ru)


@dp.message(lambda message: message.text == 'Реестры')
async def list_ru(message: types.Message):
    await message.answer(ao_ru_str)
    await message.answer(
        "Если хотите передать реестр на обслуживание в Регистратор, то пишите нам на электронную почту почту info@crc-reg.com ",
        reply_markup=greet_kb2_ru)


@dp.message(lambda message: message.text == 'Контакты')
async def contacts_ru(message: types.Message):
    await message.answer(contact_str_ru_replace + bank_str_ru, reply_markup=greet_kb2_ru)


@dp.message(lambda message: message.text == 'Порядок заполнения и предоставления анкет, распоряжений и запросов')
async def p_f_s_ru(message: types.Message):
    await message.answer("Выберите документ, который Вас интересует",
                         reply_markup=greet_kb3_ru.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Документы для проведения операций')
async def d_c_t_ru(message: types.Message):
    await message.answer("Выберите операцию, которая Вас интересует",
                         reply_markup=greet_kb4_ru.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Документы для открытия лицевого счета')
async def d_o_a_ru(message: types.Message):
    await message.answer(message.text)
    await message.answer("Если Вы не нашли нужную Вам операцию, то пишите нам на электронную почту info@crc-reg.com",
                         reply_markup=inline_kb_full5_ru)


@dp.message(lambda message: message.text == 'Документы для внесения записи в информацию ЛС о ЗЛ')
async def d_t_p_i_ru(message: types.Message):
    await message.answer(message.text)
    await message.answer("Если Вы не нашли нужную Вам операцию, то пишите нам на электронную почту info@crc-reg.com",
                         reply_markup=inline_kb_full6_ru)


@dp.message(lambda message: message.text == 'Прейскуранты')
async def price_ru(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_ru)


link_issuers = 'https://crc-reg.com/for-issuers/price-lists/'
folder_location_issuers = r''
response_issuers = requests.get(link_issuers)
soup_price_issuers_ru = BeautifulSoup(response_issuers.text, "html.parser")


@dp.callback_query(F.data == 'btn1_ru')
async def price_ru(call: types.CallbackQuery):
    find_href_issuers = soup_price_issuers_ru.select("td>a[href$='.pdf']")[0]
    filename_issuers = os.path.join(folder_location_issuers, find_href_issuers['href'].split('/')[-1])
    with open(filename_issuers, 'wb') as f:
        f.write(requests.get(urljoin(link_issuers, find_href_issuers['href'])).content)
    with open(filename_issuers, 'rb') as file1:
        doc1 = BufferedInputFile(file1.read(), filename="Прейскурант Эмитенты доп услуги_ООО_КРК_с_06.07.21.pdf")
        await call.message.answer_document(doc1, caption='Прейскурант услуг, предоставляемых Эмитенту ценных бумаг',
                                           reply_markup=greet_kb10_ru.as_markup(resize_keyboard=True))
    file1.close()


@dp.callback_query(F.data == 'btn53_ru')
async def ind_price_ru(call: types.CallbackQuery):
    find_href_issuers_2 = soup_price_issuers_ru.select("td>a[href$='.pdf']")[1]
    filename_issuers_2 = os.path.join(folder_location_issuers, find_href_issuers_2['href'].split('/')[-1])
    with open(filename_issuers_2, 'wb') as f:
        f.write(requests.get(urljoin(link_issuers, find_href_issuers_2['href'])).content)
    with open(filename_issuers_2, 'rb') as file2:
        doc2 = BufferedInputFile(file2.read(), filename="Индивидуальный Прейскурант Эмитент доп услуги_Алсиб.pdf")
        await call.message.answer_document(doc2, caption='Прейскурант услуг, предоставляемых МКПАО «Алсиб»',
                                           reply_markup=greet_kb10_ru.as_markup(resize_keyboard=True))
    file2.close()


link_shareholders = 'https://crc-reg.com/for-shareholders/price-lists/'
folder_location_shareholders = r''
response_shareholders = requests.get(link_shareholders)
soup_shareholders = BeautifulSoup(response_shareholders.text, "html.parser")


@dp.callback_query(F.data == 'btn2_ru')
async def price_ru(call: types.CallbackQuery):
    find_href_shareholders = soup_shareholders.select("a[href$='.pdf']")[1]
    filename_shareholders = os.path.join(folder_location_shareholders, find_href_shareholders['href'].split('/')[-1])
    with open(filename_shareholders, 'wb') as f:
        f.write(requests.get(urljoin(link_shareholders, find_href_shareholders['href'])).content)
    with open(filename_shareholders, 'rb') as file3:
        doc3 = BufferedInputFile(file3.read(), filename="Прейскурант Зарег. лица_ООО_КРК_с_06.07.21.pdf")
        await call.message.answer_document(doc3, caption='Прейскурант услуг, предоставляемых зарегистрированному лицу',
                                           reply_markup=greet_kb10_ru.as_markup(resize_keyboard=True))
    file3.close()


@dp.callback_query(F.data == 'btn3_ru')
async def price_ru(call: types.CallbackQuery):
    find_href_shareholders_2 = soup_shareholders.select("a[href$='.pdf']")[2]
    filename_shareholders_2 = os.path.join(folder_location_shareholders,
                                           find_href_shareholders_2['href'].split('/')[-1])
    with open(filename_shareholders_2, 'wb') as f:
        f.write(requests.get(urljoin(link_shareholders, find_href_shareholders_2['href'])).content)
    with open(filename_shareholders_2, 'rb') as file4:
        doc4 = BufferedInputFile(file4.read(), filename="Прейскурант Зарег. лица (доп. услуги)_ООО_КРК_с_06.07.21.pdf")
        await call.message.answer_document(doc4,
                                           caption='Прейскурант дополнительных услуг, предоставляемых зарегистрированному лицу',
                                           reply_markup=greet_kb10_ru.as_markup(resize_keyboard=True))
    file4.close()


@dp.callback_query(F.data == 'btn54_ru')
async def ind_price_ru(call: types.CallbackQuery):
    find_href_shareholders_3 = soup_shareholders.select("a[href$='.pdf']")[3]
    filename_shareholders_3 = os.path.join(folder_location_shareholders,
                                           find_href_shareholders_3['href'].split('/')[-1])
    with open(filename_shareholders_3, 'wb') as f:
        f.write(requests.get(urljoin(link_shareholders, find_href_shareholders_3['href'])).content)
    with open(filename_shareholders_3, 'rb') as file5:
        doc5 = BufferedInputFile(file5.read(), filename="Индивидуальный Прейскурант Зарег. лица_Алсиб.pdf")
        await call.message.answer_document(doc5, caption='Прейскурант услуг зарегистрированным лицам МКПАО «Алсиб»',
                                           reply_markup=greet_kb10_ru.as_markup(resize_keyboard=True))
    file5.close()


@dp.callback_query(F.data == 'btn55_ru')
async def ind_price_ru(call: types.CallbackQuery):
    find_href_shareholders_4 = soup_shareholders.select("a[href$='.pdf']")[4]
    filename_shareholders_4 = os.path.join(folder_location_shareholders,
                                           find_href_shareholders_4['href'].split('/')[-1])
    with open(filename_shareholders_4, 'wb') as f:
        f.write(requests.get(urljoin(link_shareholders, find_href_shareholders_4['href'])).content)
    with open(filename_shareholders_4, 'rb') as file6:
        doc6 = BufferedInputFile(file6.read(), filename="Индивидуальный Прейскурант Эмитент доп услуги_Алсиб.pdf")
        await call.message.answer_document(doc6, caption='Прейскурант доп.услуг зарегистрированным лицам МКПАО «Алсиб»',
                                           reply_markup=greet_kb10_ru.as_markup(resize_keyboard=True))
    file6.close()


@dp.message(lambda message: message.text == 'English language')
async def en_language(message: types.Message):
    await message.answer('English is selected.\n Select the desired button in the menu',
                         reply_markup=greet_kb_en.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'To the main menu')
async def main_menu_en(message: types.Message):
    await message.answer('Select the desired button in the menu',
                         reply_markup=greet_kb_en.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Go back to the section "Application forms"')
async def back_a_f_en(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_en)


@dp.message(lambda message: message.text == 'Go back to the section "Orders"')
async def back_orders_en(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_2_en)


@dp.message(lambda message: message.text == 'Go back to the section "Procedure for submitting documents"')
async def back_p_s_d_en(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_3_en)


@dp.message(
    lambda message: message.text == 'Go back to the section "Procedure for submitting documents to open a personal account"')
async def back_p_p_a_en(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_4_en)


@dp.message(
    lambda message: message.text == 'Go back to the section "Procedure for submitting documents for a transaction and for providing information"')
async def back_p_s_t_p_en(message: types.Message):
    await message.answer(message.text, reply_markup=inline_kb_full_5_en)


@dp.message(lambda message: message.text == 'Contacts')
async def contacts_en(message: types.Message):
    await message.answer(contact_str_en_replace, reply_markup=greet_kb_2_en.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Go back to the section "To the main menu"')
async def back_p_s_t_p_en(message: types.Message):
    await message.answer(message.text, reply_markup=greet_kb_en.as_markup(resize_keyboard=True))


@dp.message(
    lambda message: message.text == 'Procedure for filling out and submitting Application forms, Orders and Requests')
async def p_f_s_en(message: types.Message):
    await message.answer('Select the document you are interested in',
                         reply_markup=greet_kb_3_en.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Documents for conducting transactions')
async def d_c_t_en(message: types.Message):
    await message.answer('Select the operation that interests you',
                         reply_markup=greet_kb_4_en.as_markup(resize_keyboard=True))


@dp.message(lambda message: message.text == 'Procedure for submitting documents to open a personal account')
async def p_p_a_en(message: types.Message):
    await message.answer(message.text)
    await message.answer(
        'If you have not found the operation you need, then write to us by e - mail info @ crc - reg.com',
        reply_markup=inline_kb_full_4_en)


@dp.message(
    lambda message: message.text == 'Procedure for submitting documents for a transaction and for providing information')
async def p_s_t_p_en(message: types.Message):
    await message.answer(message.text)
    await message.answer(
        'If you have not found the operation you need, then write to us by e - mail info @ crc - reg.com',
        reply_markup=inline_kb_full_5_en)


URL_CONTACTS_en = 'https://crc-reg.com/en/about/contacts/'

r_contact_en = requests.get(URL_CONTACTS_en)
soup_contact_en = BeautifulSoup(r_contact_en.text, 'html.parser')
contacts_en = soup_contact_en.find_all('div', class_='cnt')
contact_str_en = ' '
for contact_en in contacts_en:
    contact_str_en = contact_str_en.join(contact_en.text.split(' '))
contact_str_en_replace = contact_str_en.replace('Contacts', '')


@dp.message(F.text)
async def from_file(message: types.Message):
    look_for = message.text

    wb = load_workbook(filename='file.xlsx')
    ws = wb.active

    for cell_ru in range(1, 2):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await message.answer(text=value_ru[len_ru:len_ru + 4096])
                    await message.answer(
                        "Если Вы не нашли нужную Вам анкету, то пишите нам на электронную почту info@crc-reg.com",
                        reply_markup=inline_kb_full2_ru)
            else:
                await message.answer(text=value_ru)
                await message.answer(
                    "Если Вы не нашли нужную Вам анкету, то пишите нам на электронную почту info@crc-reg.com",
                    reply_markup=inline_kb_full2_ru)

    for cell_ru in range(14, 15):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await message.answer(text=value_ru[len_ru:len_ru + 4096])
                    await message.answer(
                        "Если Вы не нашли нужное Вам Распоряжение, то пишите нам на электронную почту info@crc-reg.com",
                        reply_markup=inline_kb_full3_ru)
            else:
                await message.answer(text=value_ru)
                await message.answer(
                    "Если Вы не нашли нужное Вам Распоряжение, то пишите нам на электронную почту info@crc-reg.com",
                    reply_markup=inline_kb_full3_ru)

    for cell_ru in range(23, 24):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await message.answer(text=value_ru[len_ru:len_ru + 4096], reply_markup=inline_kb_full4_ru)
            else:
                await message.answer(text=value_ru, reply_markup=inline_kb_full4_ru)

    for cell_en in range(49, 50):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await message.answer(text=value_en[len_en:len_en + 4096])
                    await message.answer(
                        'If you have not found the Application form you need, then write to us by e - mail info @ crc - reg.com',
                        reply_markup=inline_kb_full_en)
            else:
                await message.answer(text=value_en)
                await message.answer(
                    'If you have not found the Application form you need, then write to us by e - mail info @ crc - reg.com',
                    reply_markup=inline_kb_full_en)

    for cell_en in range(58, 59):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await message.answer(text=value_en[len_en:len_en + 4096])
                    await message.answer(
                        'If you have not found the Order you need, then write to us by e-mail info@crc-reg.com',
                        reply_markup=inline_kb_full_2_en)
            else:
                await message.answer(text=value_en)
                await message.answer(
                    'If you have not found the Order you need, then write to us by e-mail info@crc-reg.com',
                    reply_markup=inline_kb_full_2_en)

    for cell_en in range(67, 68):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await message.answer(text=value_en[len_en:len_en + 4096], reply_markup=inline_kb_full_3_en)
            else:
                await message.answer(text=value_en, reply_markup=inline_kb_full_3_en)
    wb.close()


@dp.callback_query(lambda c: c.data and c.data.startswith('btn'))
async def from_file(call: types.CallbackQuery):
    look_for = call.data

    wb = load_workbook(filename='file.xlsx')
    ws = wb.active

    for cell_ru in range(2, 14):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await call.message.answer(text=value_ru[len_ru:len_ru + 4096],
                                              reply_markup=greet_kb5_ru.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_ru, reply_markup=greet_kb5_ru.as_markup(resize_keyboard=True))
    wb.close()

    for cell_ru in range(15, 23):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await call.message.answer(text=value_ru[len_ru:len_ru + 4096],
                                              reply_markup=greet_kb6_ru.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_ru, reply_markup=greet_kb6_ru.as_markup(resize_keyboard=True))
    wb.close()

    for cell_ru in range(24, 26):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await call.message.answer(text=value_ru[len_ru:len_ru + 4096],
                                              reply_markup=greet_kb7_ru.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_ru, reply_markup=greet_kb7_ru.as_markup(resize_keyboard=True))
    wb.close()

    for cell_ru in range(26, 32):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await call.message.answer(text=value_ru[len_ru:len_ru + 4096],
                                              reply_markup=greet_kb8_ru.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_ru, reply_markup=greet_kb8_ru.as_markup(resize_keyboard=True))
    wb.close()

    for cell_ru in range(32, 49):
        key_value_ru = ws.cell(row=cell_ru, column=1).value
        if key_value_ru == look_for:
            value_ru = ws.cell(row=cell_ru, column=2).value
            if len(value_ru) > 4096:
                for len_ru in range(0, len(value_ru), 4096):
                    await call.message.answer(text=value_ru[len_ru:len_ru + 4096],
                                              reply_markup=greet_kb9_ru.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_ru, reply_markup=greet_kb9_ru.as_markup(resize_keyboard=True))
    wb.close()

    for cell_en in range(50, 58):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await call.message.answer(text=value_en[len_en:len_en + 4096],
                                              reply_markup=greet_kb_5_en.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_en, reply_markup=greet_kb_5_en.as_markup(resize_keyboard=True))
    wb.close()

    for cell_en in range(59, 67):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await call.message.answer(text=value_en[len_en:len_en + 4096],
                                              reply_markup=greet_kb_6_en.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_en, reply_markup=greet_kb_6_en.as_markup(resize_keyboard=True))
    wb.close()

    for cell_en in range(68, 60):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await call.message.answer(text=value_en[len_en:len_en + 4096],
                                              reply_markup=greet_kb_7_en.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_en, reply_markup=greet_kb_7_en.as_markup(resize_keyboard=True))
    wb.close()

    for cell_en in range(70, 77):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await call.message.answer(text=value_en[len_en:len_en + 4096],
                                              reply_markup=greet_kb_8_en.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_en, reply_markup=greet_kb_8_en.as_markup(resize_keyboard=True))
    wb.close()

    for cell_en in range(77, 94):
        key_value_en = ws.cell(row=cell_en, column=1).value
        if key_value_en == look_for:
            value_en = ws.cell(row=cell_en, column=2).value
            if len(value_en) > 4096:
                for len_en in range(0, len(value_en), 4096):
                    await call.message.answer(text=value_en[len_en:len_en + 4096],
                                              reply_markup=greet_kb_9_en.as_markup(resize_keyboard=True))
            else:
                await call.message.answer(text=value_en, reply_markup=greet_kb_9_en.as_markup(resize_keyboard=True))
    wb.close()


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
